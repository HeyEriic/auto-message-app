import re
from openpyxl import load_workbook
from database.db import get_contact, insert_contacts

def read_excel_and_save_to_database(file):
  wb = load_workbook(file)
  sheet = wb.active

  for row in sheet.iter_rows(min_row=2, values_only=True):
    contact_name, _, _, contact_number, _ = row[:5]

    if contact_number is None or contact_number == "":
      print(f"Número de contato vazio para {contact_name}, pulando para o próximo contato.")
      continue

    formatted_number = format_phone_number(contact_number)
    existing_contact = get_contact(formatted_number)
    if existing_contact is None:
      insert_contacts(contact_name, formatted_number)

  wb.close()

def format_phone_number(phone_number):
  cleaned_number = re.sub(r'\D', '', phone_number)
  
  if cleaned_number.startswith('55'):
    return cleaned_number
  else:
    return '55' + cleaned_number